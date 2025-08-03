import cv2
import numpy as np
import face_recognition
import os
from datetime import date
import time
import dlib
from scipy.spatial import distance
import openpyxl
from openpyxl import Workbook, load_workbook

# URL for livestream from ESP32 CAM
ESP32_CAM_URL = "http://192.168.189.98:81/stream"
LANDMARKS_PREDICTOR_PATH = 'shape_predictor_68_face_landmarks.dat'
KNOWN_FACES_DIR = 'faces'
ATTENDANCE_EXCEL_FILE = 'attendance_excel.xlsx'

LIVENESS_TIMEOUT = 5  # Seconds to perform all actions after a blink

EYE_AR_THRESHOLD_CLOSE = 0.200
EYE_AR_THRESHOLD_OPEN = 0.255

LIP_DISTANCE_THRESHOLD = 10

HEAD_TURN_RATIO_THRESHOLD = 1.8  # Ratio of ear distances to detect a turn

FRAME_PROCESSING_FREQUENCY = 5
FACE_RECOGNITION_TOLERANCE = 0.5

liveness_tracker = {}
known_face_encodings = []
known_face_names = []
already_attended = []

try:
    predictor = dlib.shape_predictor(LANDMARKS_PREDICTOR_PATH)
except RuntimeError as e:
    print(f"Error loading landmarks predictor: {e}")
    print(f"Please ensure '{LANDMARKS_PREDICTOR_PATH}' is in the same directory.")
    exit()


def eye_aspect_ratio(eye):
    """Computes the eye aspect ratio from the eye landmarks."""
    return (distance.euclidean(eye[1], eye[5]) + distance.euclidean(eye[2], eye[4])) / (2.0 * distance.euclidean(eye[0], eye[3]))

️
def detect_real_blink(landmarks, face_id):
    """
    Detects a real blink by checking a strict 3-frame sliding window
    for an open-close-open or close-open-close pattern.
    """
    global liveness_tracker
    left_eye = np.array([(landmarks.part(i).x, landmarks.part(i).y) for i in range(36, 42)], dtype=np.int32)
    right_eye = np.array([(landmarks.part(i).x, landmarks.part(i).y) for i in range(42, 48)], dtype=np.int32)
    ear = (eye_aspect_ratio(left_eye) + eye_aspect_ratio(right_eye)) / 2.0
    tracker = liveness_tracker[face_id]

    # Determine the current state: 1 for open, 0 for closed, or -1 for indeterminate.
    current_state = -1
    if ear > EYE_AR_THRESHOLD_OPEN:
        current_state = 1  # Open
    elif ear < EYE_AR_THRESHOLD_CLOSE:
        current_state = 0  # Closed

    tracker['frame_history'].append(current_state)

    if len(tracker['frame_history']) > 3:
        tracker['frame_history'].pop(0)

    print(f"DEBUG Face {face_id} | EAR: {ear:.4f} | 3-Frame Window: {tracker['frame_history']}")    # For debugging

    if len(tracker['frame_history']) == 3:
        if tracker['frame_history'] == [1, 0, 1] or tracker['frame_history'] == [0, 1, 0]:
            print(f"  ==> SUCCESS: Blink Pattern {tracker['frame_history']} Confirmed!")
            # Clear the history to require a new sequence for the next blink.
            tracker['frame_history'] = []
            return True

    return False


def detect_lip_movement(landmarks):
    top_lip_center = np.mean([(landmarks.part(i).x, landmarks.part(i).y) for i in [50, 51, 52]], axis=0)
    bottom_lip_center = np.mean([(landmarks.part(i).x, landmarks.part(i).y) for i in [56, 57, 58]], axis=0)
    return abs(top_lip_center[1] - bottom_lip_center[1]) > LIP_DISTANCE_THRESHOLD


def detect_head_turn(landmarks):
    nose = (landmarks.part(30).x, landmarks.part(30).y)
    left_ear = (landmarks.part(0).x, landmarks.part(0).y)
    right_ear = (landmarks.part(16).x, landmarks.part(16).y)

    dist_left = distance.euclidean(nose, left_ear)
    dist_right = distance.euclidean(nose, right_ear)

    if dist_left == 0 or dist_right == 0: return False

    ratio = dist_right / dist_left
    return ratio > HEAD_TURN_RATIO_THRESHOLD or ratio < 1 / HEAD_TURN_RATIO_THRESHOLD


# ️ UPDATED STATE MACHINE INITIALIZATION ️
def liveness_state_machine(landmarks, face_id):
    global liveness_tracker
    current_time = time.time()

    if face_id not in liveness_tracker:
        liveness_tracker[face_id] = {
            'state': 'AWAITING_BLINK', 'blink_time': 0, 'has_moved_lips': False,
            'has_turned_head': False,
            'frame_history': []
        }

    tracker = liveness_tracker[face_id]
    state = tracker['state']

    if state == 'AWAITING_BLINK':
        if detect_real_blink(landmarks, face_id):
            print(f"Face {face_id}: State transition -> AWAITING_OTHER_ACTIONS")
            tracker['state'] = 'AWAITING_OTHER_ACTIONS'
            tracker['blink_time'] = current_time
        return False, "Blink to Start"

    elif state == 'AWAITING_OTHER_ACTIONS':
        if current_time - tracker['blink_time'] > LIVENESS_TIMEOUT:
            print(f"Face {face_id}: Liveness check timed out. Resetting state.")
            liveness_tracker[face_id] = {
                'state': 'AWAITING_BLINK', 'blink_time': 0, 'has_moved_lips': False,
                'has_turned_head': False, 'frame_history': []
            }
            return False, "Timeout. Blink again"

        if not tracker['has_moved_lips']:
            tracker['has_moved_lips'] = detect_lip_movement(landmarks)

        if not tracker['has_turned_head']:
            tracker['has_turned_head'] = detect_head_turn(landmarks)

        if tracker['has_moved_lips'] and tracker['has_turned_head']:
            print(f"Face {face_id}: State transition -> LIVENESS_CONFIRMED")
            tracker['state'] = 'LIVENESS_CONFIRMED'
            return True, "Liveness Confirmed"

        feedback = []
        if not tracker['has_moved_lips']: feedback.append("Move lips")
        if not tracker['has_turned_head']: feedback.append("Turn head")
        return False, " & ".join(feedback)

    elif state == 'LIVENESS_CONFIRMED':
        return True, "Liveness Confirmed"

    return False, "Initializing..."


def load_known_faces():
    print("Loading known faces...")
    for file in os.listdir(KNOWN_FACES_DIR):
        if file.endswith((".png", ".jpg", ".jpeg")):
            image_path = os.path.join(KNOWN_FACES_DIR, file)
            try:
                image = face_recognition.load_image_file(image_path)
                encodings = face_recognition.face_encodings(image)
                if encodings:
                    known_face_encodings.append(encodings[0])
                    known_face_names.append(os.path.splitext(file)[0])
            except Exception as e:
                print(f"Error loading image {file}: {e}")
    print(f"Loaded {len(known_face_names)} known faces.")


def setup_excel_sheet():
    print("Setting up Excel sheet...")
    if not os.path.exists(ATTENDANCE_EXCEL_FILE):
        wb = Workbook()
        wb.active.title = "Default"
        wb.save(ATTENDANCE_EXCEL_FILE)

    wb = load_workbook(ATTENDANCE_EXCEL_FILE)
    inp = input('Please provide a name for today\'s attendance sheet (e.g., "Lab_Session_1"): ')
    sheet_name = f"{date.today()}_{inp}"

    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
        sheet['A1'] = 'ID / Name'
        sheet['B1'] = 'Attendance Time'
    else:
        sheet = wb[sheet_name]

    for row in range(2, sheet.max_row + 1):
        name_in_sheet = sheet.cell(row=row, column=1).value
        if name_in_sheet:
            already_attended.append(name_in_sheet)
    print(f"{len(already_attended)} people have already attended: {', '.join(already_attended)}")
    return wb, sheet



if __name__ == "__main__":
    load_known_faces()
    wb, sheet = setup_excel_sheet()
    next_row = sheet.max_row + 1

    video_capture = cv2.VideoCapture(ESP32_CAM_URL)
    if not video_capture.isOpened():
        print("Error: Could not open ESP32-CAM stream.")
        exit()
    print("Successfully connected to ESP32-CAM stream. Press 'q' to quit.")

    frame_counter = 0

    while True:
        ret, frame = video_capture.read()
        if not ret:
            time.sleep(2)
            continue

        frame_counter += 1
        if frame_counter % FRAME_PROCESSING_FREQUENCY != 0:
            cv2.imshow('Face Attendance', frame)
            if cv2.waitKey(1) & 0xFF == ord('q'): break
            continue

        small_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
        gray_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2GRAY)
        face_locations = face_recognition.face_locations(gray_small_frame)

        active_face_ids = set(range(len(face_locations)))
        for inactive_id in set(liveness_tracker.keys()) - active_face_ids:
            del liveness_tracker[inactive_id]

        for i, (top, right, bottom, left) in enumerate(face_locations):
            face_id = i
            top_orig, right_orig, bottom_orig, left_orig = top * 2, right * 2, bottom * 2, left * 2

            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            face_rect = dlib.rectangle(left_orig, top_orig, right_orig, bottom_orig)
            landmarks = predictor(gray_frame, face_rect)

            is_live, feedback = liveness_state_machine(landmarks, face_id)

            if is_live:
                cv2.rectangle(frame, (left_orig, top_orig), (right_orig, bottom_orig), (0, 255, 0), 2)

                rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                face_encoding = face_recognition.face_encodings(rgb_small_frame, [(top, right, bottom, left)])[0]
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding,
                                                         FACE_RECOGNITION_TOLERANCE)
                name = "Unknown"
                if True in matches:
                    face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index]: name = known_face_names[best_match_index]

                label = f"LIVE: {name}"
                cv2.putText(frame, label, (left_orig, top_orig - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

                if name != "Unknown" and name not in already_attended:
                    sheet.cell(row=next_row, column=1, value=name)
                    sheet.cell(row=next_row, column=2, value=time.strftime("%Y-%m-%d %H:%M:%S"))
                    wb.save(ATTENDANCE_EXCEL_FILE)
                    already_attended.append(name)
                    next_row += 1
                    print(f"Attendance marked for {name}.")
            else:
                cv2.rectangle(frame, (left_orig, top_orig), (right_orig, bottom_orig), (0, 0, 255), 2)
                cv2.putText(frame, f"Action: {feedback}", (left_orig, bottom_orig + 25), cv2.FONT_HERSHEY_SIMPLEX,
                            0.6, (255, 255, 255), 2)

        cv2.imshow('Face Attendance', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'): break

    print("Closing application.")
    wb.save(ATTENDANCE_EXCEL_FILE)
    video_capture.release()
    cv2.destroyAllWindows()
